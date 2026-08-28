import os
import logging
import re
import sqlite3
import random
import csv
import io
import json
import asyncio
from datetime import datetime, timedelta
from collections import Counter
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import Application, MessageHandler, filters, ContextTypes, CommandHandler, CallbackQueryHandler
from openai import OpenAI
import matplotlib.pyplot as plt
import numpy as np
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io as io_bytes

# ---------- НАСТРОЙКИ ----------
TOKEN = "8960258146:AAEooW9g65ngBevd9lZYfJhSGA-qorb63lg"
if not TOKEN:
    raise ValueError("TOKEN не задан")

GROUP_CHAT_ID = -4462437609
DEFAULT_RESPONSIBLE = ["tunduk_dev", "tunduk_analyst"]
ADMIN_IDS = [549890508]  # Ваш ID
BOT_USERNAME = "@Jardam4y"  # ← замените на username вашего нового бота (без @)

MORNING_TIME_UTC = "03:00"
TIMEZONE_OFFSET = 6

PRIORITY_REMINDER_MINUTES = {
    "high": 5,
    "medium": 15,
    "low": 30
}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------- НАСТРОЙКА ИИ ----------
AI_API_KEY = os.environ.get("AI_BOT")
ai_client = None
if AI_API_KEY:
    try:
        ai_client = OpenAI(
            api_key=AI_API_KEY,
            base_url="https://api.deepseek.com"
        )
        logger.info("AI клиент инициализирован")
    except Exception as e:
        logger.error(f"Ошибка инициализации AI: {e}")
else:
    logger.warning("DEEPSEEK_API_KEY не задан, ИИ-функции отключены")

# ---------- БАЗА ДАННЫХ ----------
def init_db():
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS issues (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message_id INTEGER,
        chat_id INTEGER,
        author_id INTEGER,
        author_name TEXT,
        username TEXT,
        text TEXT,
        type TEXT,
        status TEXT DEFAULT 'open',
        priority TEXT DEFAULT 'low',
        tags TEXT,
        responsible TEXT,
        created_at TIMESTAMP,
        reminder_sent INTEGER DEFAULT 0,
        closed_by INTEGER,
        closed_at TIMESTAMP,
        file_id TEXT,
        file_url TEXT,
        title TEXT
    )''')
    c.execute("PRAGMA table_info(issues)")
    columns = [col[1] for col in c.fetchall()]
    for col in ["priority", "tags", "username", "responsible", "closed_by", "closed_at", "file_id", "file_url", "title"]:
        if col not in columns:
            c.execute(f"ALTER TABLE issues ADD COLUMN {col} TEXT")
    c.execute('''CREATE TABLE IF NOT EXISTS comments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        issue_id INTEGER,
        user_id INTEGER,
        user_name TEXT,
        text TEXT,
        created_at TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS rating (
        user_id INTEGER PRIMARY KEY,
        points INTEGER DEFAULT 0,
        last_updated TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        issue_id INTEGER,
        user_id INTEGER,
        action TEXT,
        old_value TEXT,
        new_value TEXT,
        created_at TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS votes (
        issue_id INTEGER,
        user_id INTEGER,
        vote INTEGER,
        PRIMARY KEY (issue_id, user_id)
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS banned_users (
        user_id INTEGER PRIMARY KEY,
        reason TEXT,
        banned_at TIMESTAMP
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS responsible_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE
    )''')
    for username in DEFAULT_RESPONSIBLE:
        c.execute("INSERT OR IGNORE INTO responsible_users (username) VALUES (?)", (username,))
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('auto_close_days', '14')")
    conn.commit()
    conn.close()
    logger.info("База данных инициализирована")

# ---------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------
def get_setting(key, default=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key=?", (key,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else default

def set_setting(key, value):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()
    conn.close()

def is_banned(user_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT 1 FROM banned_users WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row is not None

def ban_user(user_id, reason=""):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO banned_users (user_id, reason, banned_at) VALUES (?, ?, ?)",
              (user_id, reason, datetime.now()))
    conn.commit()
    conn.close()

def unban_user(user_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("DELETE FROM banned_users WHERE user_id=?", (user_id,))
    conn.commit()
    conn.close()

def add_audit_log(issue_id, user_id, action, old_val="", new_val=""):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("INSERT INTO audit_log (issue_id, user_id, action, old_value, new_value, created_at) VALUES (?, ?, ?, ?, ?, ?)",
              (issue_id, user_id, action, old_val, new_val, datetime.now()))
    conn.commit()
    conn.close()

def add_vote(issue_id, user_id, vote):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO votes (issue_id, user_id, vote) VALUES (?, ?, ?)",
              (issue_id, user_id, vote))
    conn.commit()
    conn.close()
    c = conn.cursor()
    c.execute("SELECT author_id FROM issues WHERE id=?", (issue_id,))
    row = c.fetchone()
    if row:
        add_points(row[0], vote)

def get_votes(issue_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT SUM(vote) FROM votes WHERE issue_id=?", (issue_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row and row[0] else 0

def add_points(user_id, points):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("INSERT INTO rating (user_id, points, last_updated) VALUES (?, ?, ?) "
              "ON CONFLICT(user_id) DO UPDATE SET points = points + ?, last_updated = ?",
              (user_id, points, datetime.now(), points, datetime.now()))
    conn.commit()
    conn.close()

def get_points(user_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT points FROM rating WHERE user_id=?", (user_id,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else 0

def get_top_users(limit=10):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT user_id, points FROM rating ORDER BY points DESC LIMIT ?", (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_responsible_list():
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT username FROM responsible_users ORDER BY username")
    rows = c.fetchall()
    conn.close()
    return [row[0] for row in rows]

def add_responsible(username):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO responsible_users (username) VALUES (?)", (username,))
    conn.commit()
    conn.close()

def remove_responsible(username):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("DELETE FROM responsible_users WHERE username=?", (username,))
    conn.commit()
    conn.close()

def get_issue_by_id(issue_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT id, status, text, author_name, responsible, priority FROM issues WHERE id=?", (issue_id,))
    row = c.fetchone()
    conn.close()
    return row

def get_issue_by_reply(reply_to_message_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT id, status FROM issues WHERE message_id=?", (reply_to_message_id,))
    row = c.fetchone()
    conn.close()
    return row

def add_issue(message, issue_type, tags, responsible, priority, file_id="", file_url="", title=""):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute('''INSERT INTO issues 
        (message_id, chat_id, author_id, author_name, username, text, type, tags, responsible, priority, created_at, file_id, file_url, title)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (message.message_id, message.chat_id, message.from_user.id,
         message.from_user.full_name, message.from_user.username or "",
         message.text, issue_type, tags, responsible, priority, datetime.now(),
         file_id, file_url, title))
    issue_id = c.lastrowid
    conn.commit()
    conn.close()
    add_audit_log(issue_id, message.from_user.id, "create", "", f"type={issue_type}, priority={priority}")
    logger.info(f"Задача #{issue_id} создана")
    return issue_id

def update_priority(issue_id, priority, user_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT priority FROM issues WHERE id=?", (issue_id,))
    old = c.fetchone()
    c.execute("UPDATE issues SET priority=? WHERE id=?", (priority, issue_id))
    conn.commit()
    conn.close()
    if old and old[0] != priority:
        add_audit_log(issue_id, user_id, "change_priority", old[0], priority)

def close_issue(issue_id, closer_id=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT status FROM issues WHERE id=?", (issue_id,))
    old = c.fetchone()
    if old and old[0] == 'closed':
        conn.close()
        return
    if closer_id:
        c.execute("UPDATE issues SET status='closed', closed_by=?, closed_at=? WHERE id=?", (closer_id, datetime.now(), issue_id))
    else:
        c.execute("UPDATE issues SET status='closed', closed_at=? WHERE id=?", (datetime.now(), issue_id))
    conn.commit()
    conn.close()
    add_audit_log(issue_id, closer_id or 0, "close", old[0] if old else "", "closed")
    if closer_id:
        add_points(closer_id, 2)

def reopen_issue(issue_id, user_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT status FROM issues WHERE id=?", (issue_id,))
    old = c.fetchone()
    c.execute("UPDATE issues SET status='open', reminder_sent=0 WHERE id=?", (issue_id,))
    conn.commit()
    conn.close()
    add_audit_log(issue_id, user_id, "reopen", old[0] if old else "", "open")

def is_issue_resolved(issue_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT status FROM issues WHERE id=?", (issue_id,))
    row = c.fetchone()
    conn.close()
    return row and row[0] == 'closed'

def mark_reminder_sent(issue_id):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("UPDATE issues SET reminder_sent=1 WHERE id=?", (issue_id,))
    conn.commit()
    conn.close()

def add_comment(issue_id, user_id, user_name, text):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute('''INSERT INTO comments (issue_id, user_id, user_name, text, created_at)
                 VALUES (?, ?, ?, ?, ?)''',
              (issue_id, user_id, user_name, text, datetime.now()))
    conn.commit()
    conn.close()
    add_audit_log(issue_id, user_id, "comment", "", text)

def extract_tags(text):
    return re.findall(r'#\w+', text)

def extract_mentions(text):
    return re.findall(r'@(\w+)', text)

def detect_priority(text):
    text_lower = text.lower()
    if re.search(r'критичн|срочн|high|critical', text_lower):
        return 'high'
    elif re.search(r'важн|medium|normal', text_lower):
        return 'medium'
    else:
        return 'low'

def get_all_tags():
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT DISTINCT tags FROM issues WHERE tags IS NOT NULL AND tags != ''")
    rows = c.fetchall()
    conn.close()
    tags_set = set()
    for row in rows:
        if row[0]:
            for tag in row[0].split(','):
                if tag:
                    tags_set.add(tag.strip())
    return sorted(tags_set)

def get_stats_by_tag(tag, days=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = "SELECT type, status FROM issues WHERE tags LIKE ?"
    params = [f"%{tag}%"]
    if days:
        cutoff = datetime.now() - timedelta(days=days)
        query += " AND created_at >= ?"
        params.append(cutoff)
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    stats = {"bug": {"total": 0, "closed": 0}, "suggestion": {"total": 0, "closed": 0}}
    for row in rows:
        t = row[0]
        status = row[1]
        stats[t]["total"] += 1
        if status == "closed":
            stats[t]["closed"] += 1
    return stats

def get_user_stats(user_id, days=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = "SELECT type, status FROM issues WHERE author_id=?"
    params = [user_id]
    if days:
        cutoff = datetime.now() - timedelta(days=days)
        query += " AND created_at >= ?"
        params.append(cutoff)
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    stats = {"bug": {"total": 0, "closed": 0}, "suggestion": {"total": 0, "closed": 0}}
    for row in rows:
        t = row[0]
        status = row[1]
        stats[t]["total"] += 1
        if status == "closed":
            stats[t]["closed"] += 1
    return stats

def get_user_stats_by_username(username, days=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = "SELECT type, status FROM issues WHERE username=?"
    params = [username]
    if days:
        cutoff = datetime.now() - timedelta(days=days)
        query += " AND created_at >= ?"
        params.append(cutoff)
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    stats = {"bug": {"total": 0, "closed": 0}, "suggestion": {"total": 0, "closed": 0}}
    for row in rows:
        t = row[0]
        status = row[1]
        stats[t]["total"] += 1
        if status == "closed":
            stats[t]["closed"] += 1
    return stats

def get_stats_responsible(username, days=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = "SELECT id, status, type FROM issues WHERE responsible LIKE ?"
    params = [f"%{username}%"]
    if days:
        cutoff = datetime.now() - timedelta(days=days)
        query += " AND created_at >= ?"
        params.append(cutoff)
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    total = len(rows)
    closed = sum(1 for r in rows if r[1] == 'closed')
    bugs = sum(1 for r in rows if r[2] == 'bug')
    suggestions = sum(1 for r in rows if r[2] == 'suggestion')
    return total, closed, bugs, suggestions

def search_issues(text_query):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT id, author_name, text, type, status, tags, created_at FROM issues WHERE text LIKE ? ORDER BY created_at DESC LIMIT 20", (f"%{text_query}%",))
    rows = c.fetchall()
    conn.close()
    return rows

def generate_export(days=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = "SELECT id, author_name, username, text, type, status, priority, tags, responsible, created_at FROM issues"
    params = []
    if days:
        cutoff = datetime.now() - timedelta(days=days)
        query += " WHERE created_at >= ?"
        params.append(cutoff)
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    return rows

def generate_weekly_report():
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    week_ago = datetime.now() - timedelta(days=7)
    c.execute("SELECT COUNT(*) FROM issues WHERE created_at >= ?", (week_ago,))
    total_created = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM issues WHERE status='closed' AND created_at >= ?", (week_ago,))
    total_closed = c.fetchone()[0]
    c.execute("SELECT type, COUNT(*) FROM issues WHERE created_at >= ? GROUP BY type", (week_ago,))
    type_counts = c.fetchall()
    c.execute("SELECT priority, COUNT(*) FROM issues WHERE created_at >= ? GROUP BY priority", (week_ago,))
    priority_counts = c.fetchall()
    conn.close()
    return total_created, total_closed, type_counts, priority_counts

def get_open_tasks_text(limit=10, priority=None, tag=None, days=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = "SELECT id, author_name, text, type, priority, tags, created_at FROM issues WHERE status='open'"
    params = []
    if priority:
        query += " AND priority=?"
        params.append(priority)
    if tag:
        query += " AND tags LIKE ?"
        params.append(f"%{tag}%")
    if days:
        cutoff = datetime.now() - timedelta(days=days)
        query += " AND created_at >= ?"
        params.append(cutoff)
    query += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()
    return rows

def is_greeting_or_question(text):
    text_lower = text.lower()
    patterns = [r'\bпривет\b', r'\bздравствуй[те]?\b', r'\bсалам\b', r'\bhello\b', r'\bhi\b',
                r'кто ты', r'откуда', r'зачем', r'для чего', r'что ты умеешь', r'расскажи о себе']
    for p in patterns:
        if re.search(p, text_lower):
            return True
    return False

def generate_excel_report(from_date=None, to_date=None):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    query = """
        SELECT id, title, text, author_name, responsible, created_at, closed_at, priority, tags
        FROM issues
        WHERE status='closed'
    """
    params = []
    if from_date:
        query += " AND closed_at >= ?"
        params.append(from_date)
    if to_date:
        query += " AND closed_at <= ?"
        params.append(to_date)
    query += " ORDER BY responsible, closed_at"
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Решённые задачи"
    headers = ["ID", "Заголовок", "Текст", "Автор", "Ответственный", "Создано", "Закрыто", "Приоритет", "Теги"]
    ws.append(headers)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([
            row[0], row[1] or "", row[2] or "", row[3] or "", row[4] or "",
            row[5] if row[5] else "",
            row[6] if row[6] else "",
            row[7] or "", row[8] or ""
        ])
    for col in range(1, len(headers)+1):
        col_letter = get_column_letter(col)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col_letter].width = adjusted_width

    ws2 = wb.create_sheet("Сводка по ответственным")
    ws2.append(["Ответственный", "Количество закрытых задач"])
    counter = Counter()
    for row in rows:
        resp = row[4] if row[4] else "Не назначен"
        counter[resp] += 1
    for resp, count in counter.most_common():
        ws2.append([resp, count])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ---------- АНЕКДОТЫ ----------
JOKES = [
    "— Почему программисты путают Хэллоуин и Рождество? — Потому что 31 Oct = 25 Dec.",
    "— Сколько программистов нужно, чтобы заменить лампочку? — Ни одного, это аппаратная проблема.",
    "— Что говорит программист, когда теряет работу? — 'Я оказался в бесконечном цикле без выхода.'",
    "— Чем отличается программист от электрика? — Электрик не пытается починить выключатель в розетке с помощью блока try...catch.",
    "— Бета-тестер: 'Я нашёл баг!' Разработчик: 'Это не баг, это фича.'",
    "— Почему у программистов всегда холодно? — Потому что они работают в среде с постоянным 'wind' (ветер).",
    "— Что делает программист, когда приходит к врачу? — Проверяет версию прошивки.",
    "— Как отличить хорошего программиста от плохого? — Хороший пишет код, который работает; плохой — код, который красиво написан.",
    "— Почему программисты любят темный режим? — Потому что свет привлекает баги.",
    "— Что такое идеальный код? — Тот, который работает без ошибок, но никто не знает, как он работает.",
    "— Бермет тестирует новый релиз. Говорит: 'Всё работает'. Через минуту — 'Ой, а это что? Баг?'. Так она нашла 10 багов за 5 минут.",
    "— Толгонай говорит разработчику: 'Ты исправил баг?'. Разработчик: 'Да'. Толгонай: 'А почему тогда у меня опять не работает?'. Разработчик: 'Это ты не тот кнопку нажала'. ",
    "— Бекзат пишет код так быстро, что компилятор просит его остановиться и дать ему отдохнуть.",
    "— Адил пришёл на работу за 2 часа до дедлайна. Коллеги: 'Ты успеешь?'. Адил: 'Я не успею, но я попрошу у босса ещё час'. И попросил. И получил.",
    "— Нурзада говорит: 'Я напишу документацию'. Все засмеялись. Но она написала. И она была такой подробной, что даже баги начали сами собой исправляться.",
    "— Айдай нарисовала макет. Разработчик: 'Это невозможно реализовать'. Айдай: 'Тогда я перерисую'. Через час — 'А так можно?'. И так 5 раз. В итоге сделали как в первом варианте.",
    "— Улан: 'Я подниму сервер за минуту'. Все удивились. Через минуту сервер действительно работал. Правда, через две минуты он упал. Но Улан сказал: 'Это была нагрузочная проверка'.",
    "— Калыс анализирует данные: 'У нас 95% багов происходит в пятницу вечером'. Коллеги: 'И что?'. Калыс: 'Значит, в пятницу вечером надо не работать, а отдыхать'.",
    "— Жанара: 'Коллеги, давайте обсудим задачи'. Все молчат. Жанара: 'Тогда я напишу в чат'. И написала. И все ответили. Вот что значит коммуникация!",
    "— Ыкыбал оптимизирует код: 'Я убрал 100 строк и теперь всё работает в 2 раза быстрее'. Коллеги: 'А как?'. Ыкыбал: 'Я просто удалил комментарии'."
]

def get_random_joke():
    return random.choice(JOKES)

# ---------- ФУНКЦИИ AI ----------
async def analyze_with_ai(text: str):
    if ai_client is None:
        return {"type": "other", "reason": "AI недоступен"}
    try:
        response = ai_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": """
                Ты — классификатор сообщений для IT-команды.
                Проанализируй сообщение и определи его тип:
                - "bug": если пользователь сообщает об ошибке, проблеме или неработающей функции.
                - "suggestion": если пользователь вносит предложение по улучшению, новую идею.
                - "other": если сообщение не относится к первым двум категориям.
                Ответь ТОЛЬКО в формате JSON: {"type": "тип", "reason": "краткая причина"}.
                """},
                {"role": "user", "content": text}
            ],
            response_format={"type": "json_object"}
        )
        return json.loads(response.choices[0].message.content)
    except Exception as e:
        logger.error(f"AI ошибка: {e}")
        return {"type": "other", "reason": "Ошибка анализа"}

async def answer_with_ai(question: str) -> str:
    if ai_client is None:
        return "❌ AI не настроен."
    try:
        if len(question) > 500:
            return "⚠️ Вопрос слишком длинный (макс. 500 символов)."
        response = ai_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "Ты — полезный и информативный ассистент."},
                {"role": "user", "content": question}
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        logger.error(f"AI ответ ошибка: {e}")
        return "❌ Ошибка при обработке вопроса."

async def generate_title(text: str) -> str:
    if ai_client is None:
        return text[:50] + "..."
    try:
        response = ai_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "Сформулируй краткий заголовок для задачи (максимум 10 слов) на основе текста. Ответь только заголовком, без пояснений."},
                {"role": "user", "content": text}
            ]
        )
        return response.choices[0].message.content.strip()
    except Exception:
        return text[:50] + "..."

async def find_similar_issues(text: str, limit=3):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT id, text, author_name FROM issues WHERE status='open' ORDER BY created_at DESC LIMIT 20")
    rows = c.fetchall()
    conn.close()
    if not rows:
        return []
    words = set(re.findall(r'\w+', text.lower()))
    scores = []
    for row in rows:
        issue_id, issue_text, author = row
        common = len(words & set(re.findall(r'\w+', issue_text.lower())))
        if common > 1:
            scores.append((common, issue_id, issue_text[:50], author))
    scores.sort(reverse=True)
    return scores[:limit]

async def auto_tagging(text: str):
    if ai_client is None:
        return []
    try:
        response = ai_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "Ты — система автоматического тегирования. Проанализируй текст и предложи 2-5 хэштегов (начинающихся с #) через запятую. Ответь только тегами, без пояснений."},
                {"role": "user", "content": text}
            ]
        )
        tags = response.choices[0].message.content.strip()
        return re.findall(r'#\w+', tags)
    except Exception:
        return []

async def smart_priority(text: str):
    if ai_client is None:
        return detect_priority(text)
    try:
        response = ai_client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "Определи приоритет задачи: high (критично), medium (важно), low (обычное). Ответь только одним словом."},
                {"role": "user", "content": text}
            ]
        )
        p = response.choices[0].message.content.strip().lower()
        if p in ['high', 'medium', 'low']:
            return p
        return detect_priority(text)
    except Exception:
        return detect_priority(text)

# ---------- СОЗДАНИЕ ЗАДАЧИ ----------
async def create_issue_from_message(msg, issue_type, context, responsible=None):
    text = msg.text
    title = await generate_title(text)
    tags_list = extract_tags(text)
    tags_str = ",".join(tags_list) if tags_list else ""

    if ai_client and not tags_list:
        ai_tags = await auto_tagging(text)
        if ai_tags:
            tags_str = ",".join(ai_tags)
            tags_list = ai_tags

    priority = await smart_priority(text)

    if responsible is None:
        mentions = extract_mentions(text)
        if mentions:
            responsible = mentions
        else:
            responsible = DEFAULT_RESPONSIBLE

    responsible_str = ",".join(responsible)
    file_id = ""
    file_url = ""
    if msg.document:
        file_id = msg.document.file_id
        file_url = msg.document.file_name or "документ"
    elif msg.photo:
        file_id = msg.photo[-1].file_id
        file_url = "фото"

    issue_id = add_issue(msg, issue_type, tags_str, responsible_str, priority, file_id, file_url, title)

    responsible_mentions = " ".join([f"@{u}" for u in responsible])
    tag_text = f"🏷️ Теги: {', '.join(tags_list)}" if tags_list else ""
    priority_emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(priority, "")
    minutes = PRIORITY_REMINDER_MINUTES[priority]
    file_text = f"📎 Вложение: {file_url}" if file_url else ""

    similar = await find_similar_issues(text, limit=3)
    similar_text = ""
    if similar:
        similar_text = "\n⚠️ Похожие открытые задачи:\n"
        for _, sid, stext, sauthor in similar:
            similar_text += f"  #{sid} ({sauthor}): {stext}...\n"

    vote_buttons = []
    if issue_type == "suggestion":
        vote_buttons = [
            InlineKeyboardButton("👍", callback_data=f"vote_{issue_id}_1"),
            InlineKeyboardButton("👎", callback_data=f"vote_{issue_id}_-1")
        ]

    await msg.reply_text(
        f"🔔 Зарегистрирован {issue_type} #{issue_id}\n"
        f"Заголовок: {title}\n"
        f"Автор: {msg.from_user.full_name}\n"
        f"Текст: {text[:100]}...\n"
        f"{tag_text}\n"
        f"Приоритет: {priority_emoji} {priority}\n"
        f"Ответственные: {responsible_mentions}\n"
        f"⏳ Напоминание через {minutes} мин.\n"
        f"{file_text}\n"
        f"{similar_text}",
        reply_markup=InlineKeyboardMarkup([vote_buttons]) if vote_buttons else None
    )

    if priority == "high":
        for user in responsible:
            try:
                await context.bot.send_message(
                    chat_id=user,
                    text=f"🚨 Критичная задача #{issue_id}!\n{title}\n{text[:200]}...\nОтветственные: {responsible_mentions}"
                )
            except Exception:
                pass

    priority_keyboard = [
        [
            InlineKeyboardButton("🔴 Критичный", callback_data=f"set_priority_{issue_id}_high"),
            InlineKeyboardButton("🟡 Важный", callback_data=f"set_priority_{issue_id}_medium"),
            InlineKeyboardButton("🟢 Обычный", callback_data=f"set_priority_{issue_id}_low")
        ],
        [InlineKeyboardButton("⏩ Пропустить", callback_data=f"skip_priority_{issue_id}")]
    ]
    await msg.reply_text(
        "Выберите приоритет задачи или нажмите 'Пропустить':",
        reply_markup=InlineKeyboardMarkup(priority_keyboard)
    )

    if context.job_queue:
        delay = minutes * 60
        job = context.job_queue.run_once(send_reminder, when=delay, data={"issue_id": issue_id, "chat_id": msg.chat_id})
        if context.bot_data.get('reminder_jobs') is None:
            context.bot_data['reminder_jobs'] = {}
        context.bot_data['reminder_jobs'][issue_id] = job

    add_points(msg.from_user.id, 1)
    return issue_id

# ---------- ОБРАБОТЧИКИ КНОПОК ----------
async def priority_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id
    parts = data.split('_')
    if len(parts) < 3:
        await query.edit_message_text("❌ Неверный формат.")
        return
    action = parts[1]
    issue_id = int(parts[2])
    issue = get_issue_by_id(issue_id)
    if not issue:
        await query.edit_message_text("❌ Задача не найдена.")
        return
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT author_id FROM issues WHERE id=?", (issue_id,))
    row = c.fetchone()
    conn.close()
    author_id = row[0] if row else None
    if user_id != author_id and user_id not in ADMIN_IDS:
        await query.edit_message_text("⛔ Нет прав.")
        return
    if action == 'priority':
        new_priority = parts[3]
        update_priority(issue_id, new_priority, user_id)
        emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(new_priority, "")
        await query.edit_message_text(f"✅ Приоритет задачи #{issue_id} изменён на {emoji} {new_priority}.")
        if context.job_queue:
            jobs = context.bot_data.get('reminder_jobs', {})
            old = jobs.get(issue_id)
            if old:
                old.schedule_removal()
            minutes = PRIORITY_REMINDER_MINUTES[new_priority]
            new_job = context.job_queue.run_once(send_reminder, when=minutes*60, data={"issue_id": issue_id, "chat_id": GROUP_CHAT_ID})
            jobs[issue_id] = new_job
    elif action == 'skip':
        await query.edit_message_text(f"⏩ Приоритет задачи #{issue_id} оставлен как авто.")

async def confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if 'pending_issue' not in context.user_data:
        await query.edit_message_text("❌ Запрос устарел.")
        return
    pending = context.user_data['pending_issue']
    if query.from_user.id != pending['message'].from_user.id:
        await query.edit_message_text("⛔ Только автор может подтвердить.")
        return
    if query.data == "confirm_yes":
        await show_responsible_selection(context, query, pending, pending['issue_type'])
    else:
        await query.edit_message_text("❌ Создание отменено.")
        del context.user_data['pending_issue']

async def show_responsible_selection(context, query, pending, issue_type):
    responsible_list = get_responsible_list()
    if not responsible_list:
        responsible_list = DEFAULT_RESPONSIBLE
        for username in responsible_list:
            add_responsible(username)
    buttons = []
    row = []
    for i, username in enumerate(responsible_list):
        row.append(InlineKeyboardButton(f"@{username}", callback_data=f"resp_{username}"))
        if len(row) == 3:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton("✏️ Другой (ввести @username)", callback_data="resp_other")])
    buttons.append([InlineKeyboardButton("⏩ Пропустить", callback_data="resp_skip")])
    context.user_data['pending_responsible'] = {'pending': pending, 'issue_type': issue_type, 'query': query}
    await query.edit_message_text(
        "Выберите ответственного за задачу:",
        reply_markup=InlineKeyboardMarkup(buttons)
    )

async def responsible_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id
    if 'pending_responsible' not in context.user_data:
        await query.edit_message_text("❌ Сессия истекла.")
        return
    pending_data = context.user_data['pending_responsible']
    pending = pending_data['pending']
    issue_type = pending_data['issue_type']
    if pending['message'].from_user.id != user_id:
        await query.edit_message_text("⛔ Только автор может выбирать.")
        return
    if data == "resp_skip":
        msg = pending['message']
        text = msg.text
        mentions = extract_mentions(text)
        if mentions:
            responsible = mentions
        else:
            responsible_list = get_responsible_list()
            responsible = [responsible_list[0]] if responsible_list else DEFAULT_RESPONSIBLE
        await query.edit_message_text("✅ Создаю задачу...")
        await create_issue_from_message(msg, issue_type, context, responsible)
        del context.user_data['pending_responsible']
        del context.user_data['pending_issue']
        return
    elif data == "resp_other":
        await query.edit_message_text("✏️ Напишите в ответ на это сообщение @username ответственного.")
        context.user_data['awaiting_responsible'] = True
        context.user_data['pending_for_responsible'] = {'pending': pending, 'issue_type': issue_type, 'query': query}
        return
    elif data.startswith("resp_"):
        username = data.split("_")[1]
        responsible = [username]
        msg = pending['message']
        await query.edit_message_text(f"✅ Выбран @{username}. Создаю задачу...")
        await create_issue_from_message(msg, issue_type, context, responsible)
        del context.user_data['pending_responsible']
        del context.user_data['pending_issue']
        return

async def handle_manual_responsible(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.user_data.get('awaiting_responsible'):
        return
    msg = update.message
    if not msg or not msg.text or msg.chat_id != GROUP_CHAT_ID:
        return
    text = msg.text.strip()
    mentions = extract_mentions(text)
    if not mentions:
        await msg.reply_text("⚠️ Не найдено @username. Попробуйте ещё раз.")
        return
    data = context.user_data.get('pending_for_responsible')
    if not data:
        return
    pending = data['pending']
    issue_type = data['issue_type']
    await msg.reply_text(f"✅ Выбраны: {' '.join([f'@{u}' for u in mentions])}. Создаю задачу...")
    await create_issue_from_message(pending['message'], issue_type, context, mentions)
    del context.user_data['awaiting_responsible']
    del context.user_data['pending_for_responsible']
    del context.user_data['pending_issue']

async def vote_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    parts = data.split('_')
    if len(parts) < 3:
        return
    issue_id = int(parts[1])
    vote = int(parts[2])
    user_id = query.from_user.id
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT 1 FROM votes WHERE issue_id=? AND user_id=?", (issue_id, user_id))
    if c.fetchone():
        await query.edit_message_text("❌ Вы уже голосовали.")
        conn.close()
        return
    conn.close()
    add_vote(issue_id, user_id, vote)
    total = get_votes(issue_id)
    await query.edit_message_text(f"✅ Голос учтён! Рейтинг: {total}")

# ---------- ПРИВЕТСТВИЕ НОВЫХ ----------
async def greet_new_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    for member in update.message.new_chat_members:
        if member.id == context.bot.id:
            continue
        await update.message.reply_text(
            f"👋 Добро пожаловать, {member.full_name}!\n\n"
            "Для справки используйте /help."
        )

# ---------- УТРЕННЕЕ ПРИВЕТСТВИЕ ----------
async def morning_greeting(context: ContextTypes.DEFAULT_TYPE):
    await context.bot.send_message(
        chat_id=GROUP_CHAT_ID,
        text="🌞 Доброе утро, коллеги! Желаем продуктивного дня!"
    )
    tasks = get_open_tasks_text(limit=10)
    if tasks:
        response = "📋 Открытые задачи:\n"
        for row in tasks:
            issue_id, author, text, issue_type, priority, tags, created = row
            emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(priority, "")
            created_str = created[:16] if isinstance(created, str) else created.strftime("%Y-%m-%d %H:%M")
            response += f"#{issue_id} {issue_type} {emoji} от {author} ({created_str}): {text[:50]}...\n"
        await context.bot.send_message(chat_id=GROUP_CHAT_ID, text=response)
    else:
        await context.bot.send_message(chat_id=GROUP_CHAT_ID, text="🎉 Нет открытых задач!")

# ---------- ОСНОВНОЙ ОБРАБОТЧИК ----------
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg or not msg.text or msg.chat_id != GROUP_CHAT_ID:
        return
    text = msg.text

    if is_banned(msg.from_user.id):
        await msg.reply_text("⛔ Вы забанены.")
        return

    if msg.reply_to_message:
        issue = get_issue_by_reply(msg.reply_to_message.message_id)
        if issue and issue[1] == 'open':
            close_issue(issue[0], msg.from_user.id)
            await msg.reply_text(f"✅ Задача #{issue[0]} закрыта.")
        return

    if text.startswith('/'):
        return

    if f"@{BOT_USERNAME}" in text or is_greeting_or_question(text):
        await msg.reply_text(
            "👋 Привет! Я бот для регистрации багов и предложений.\n"
            "Используйте /help для списка команд."
        )
        return

    if ai_client is not None:
        analysis = await analyze_with_ai(text)
        issue_type = analysis.get("type") if analysis else None
        if issue_type in ["bug", "suggestion"]:
            logger.info(f"AI определил как {issue_type}: {analysis.get('reason', '')}")
            neg_phrases = ["не баг", "не ошибка", "это не баг", "не предложение", "не улучшение"]
            if any(phrase in text.lower() for phrase in neg_phrases):
                return
            context.user_data['pending_issue'] = {'message': msg, 'text': text, 'issue_type': issue_type}
            keyboard = [[
                InlineKeyboardButton("✅ Да, создать", callback_data="confirm_yes"),
                InlineKeyboardButton("❌ Нет, отменить", callback_data="confirm_no")
            ]]
            await msg.reply_text(
                f"Я определил это как '{issue_type}'. Создать задачу?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return

    bug_pattern = r'\b(баг|ошибка|bug|не работает|глюк)\b'
    suggest_pattern = r'\b(предложени|улучшени|suggest|идея|хотелось бы)\b'
    issue_type = None
    if re.search(bug_pattern, text, re.IGNORECASE):
        issue_type = "bug"
    elif re.search(suggest_pattern, text, re.IGNORECASE):
        issue_type = "suggestion"
    else:
        return

    neg_phrases = ["не баг", "не ошибка", "это не баг", "не предложение", "не улучшение", "поясню", "объясню", "просто хочу сказать", "к слову", "кстати"]
    if any(phrase in text.lower() for phrase in neg_phrases):
        return

    context.user_data['pending_issue'] = {'message': msg, 'text': text, 'issue_type': issue_type}
    keyboard = [[
        InlineKeyboardButton("✅ Да, создать", callback_data="confirm_yes"),
        InlineKeyboardButton("❌ Нет, отменить", callback_data="confirm_no")
    ]]
    await msg.reply_text(
        f"Вы упомянули '{issue_type}'. Создать задачу?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

# ---------- НАПОМИНАНИЕ ----------
async def send_reminder(context: ContextTypes.DEFAULT_TYPE):
    data = context.job.data
    issue_id = data["issue_id"]
    chat_id = data["chat_id"]
    if is_issue_resolved(issue_id):
        return
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT reminder_sent, responsible FROM issues WHERE id=?", (issue_id,))
    row = c.fetchone()
    if row and row[0] == 1:
        conn.close()
        return
    conn.close()
    mark_reminder_sent(issue_id)
    resp = row[1] if row and row[1] else ""
    mentions = ""
    if resp:
        usernames = [u.strip() for u in resp.split(',') if u.strip()]
        if usernames:
            mentions = " " + " ".join([f"@{u}" for u in usernames])
    await context.bot.send_message(
        chat_id=chat_id,
        text=f"⚠️ Напоминание! Задача #{issue_id} без ответа.{mentions}\nПросьба ответить."
    )

# ---------- КОМАНДЫ ----------
async def bug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg or not msg.text or msg.chat_id != GROUP_CHAT_ID:
        return
    text = msg.text.replace('/bug', '').strip()
    if not text:
        await msg.reply_text("Укажите описание: /bug описание")
        return
    msg.text = text
    await create_issue_from_message(msg, "bug", context)
    msg.text = "/bug " + text

async def suggest_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg or not msg.text or msg.chat_id != GROUP_CHAT_ID:
        return
    text = msg.text.replace('/suggest', '').strip()
    if not text:
        await msg.reply_text("Укажите описание: /suggest описание")
        return
    msg.text = text
    await create_issue_from_message(msg, "suggestion", context)
    msg.text = "/suggest " + text

async def priority_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if len(context.args) < 2:
        await msg.reply_text("Использование: /priority <id> <high|medium|low>")
        return
    try:
        issue_id = int(context.args[0])
    except ValueError:
        await msg.reply_text("Некорректный ID.")
        return
    pr = context.args[1].lower()
    if pr not in ['high', 'medium', 'low']:
        await msg.reply_text("Приоритет: high, medium или low")
        return
    issue = get_issue_by_id(issue_id)
    if not issue:
        await msg.reply_text(f"Задача #{issue_id} не найдена.")
        return
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT author_id FROM issues WHERE id=?", (issue_id,))
    row = c.fetchone()
    conn.close()
    author_id = row[0] if row else None
    user_id = update.effective_user.id
    if user_id != author_id and user_id not in ADMIN_IDS:
        await msg.reply_text("⛔ Нет прав.")
        return
    update_priority(issue_id, pr, user_id)
    emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(pr, "")
    await msg.reply_text(f"✅ Приоритет задачи #{issue_id} изменён на {emoji} {pr}.")
    if context.job_queue:
        jobs = context.bot_data.get('reminder_jobs', {})
        old = jobs.get(issue_id)
        if old:
            old.schedule_removal()
        minutes = PRIORITY_REMINDER_MINUTES[pr]
        new_job = context.job_queue.run_once(send_reminder, when=minutes*60, data={"issue_id": issue_id, "chat_id": GROUP_CHAT_ID})
        jobs[issue_id] = new_job

async def ask_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("❓ Напишите вопрос после команды: /ask ваш вопрос")
        return
    question = " ".join(context.args)
    if len(question) > 500:
        await update.message.reply_text("⚠️ Вопрос слишком длинный (макс. 500 символов).")
        return
    await update.message.reply_text("🤔 Думаю...")
    answer = await answer_with_ai(question)
    cleaned = re.sub(r'\*\*(.*?)\*\*', r'\1', answer)
    cleaned = re.sub(r'\*(.*?)\*', r'\1', cleaned)
    cleaned = re.sub(r'_(.*?)_', r'\1', cleaned)
    cleaned = re.sub(r'#{1,6}\s?', '', cleaned)
    cleaned = re.sub(r'`(.*?)`', r'\1', cleaned)
    cleaned = re.sub(r'\[(.*?)\]\(.*?\)', r'\1', cleaned)
    cleaned = cleaned.replace('*', '')
    await update.message.reply_text(cleaned)

async def my_stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    username = update.effective_user.username or "без юзернейма"
    days = None
    if context.args and context.args[0].lower() in ["week", "7"]:
        days = 7
    elif context.args and context.args[0].lower() in ["month", "30"]:
        days = 30
    stats = get_user_stats(user_id, days)
    period = "за всё время" if not days else f"за последние {days} дней"
    response = f"📊 Статистика @{username} {period}:\n"
    response += f"🐛 Баги: {stats['bug']['total']} создано, {stats['bug']['closed']} закрыто\n"
    response += f"💡 Предложения: {stats['suggestion']['total']} создано, {stats['suggestion']['closed']} закрыто"
    await update.message.reply_text(response)

async def stats_username_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Укажите юзернейм: /stats @username")
        return
    target = context.args[0].lstrip('@')
    days = None
    if len(context.args) > 1 and context.args[1].lower() in ["week", "7"]:
        days = 7
    elif len(context.args) > 1 and context.args[1].lower() in ["month", "30"]:
        days = 30
    stats = get_user_stats_by_username(target, days)
    if stats["bug"]["total"] == 0 and stats["suggestion"]["total"] == 0:
        await update.message.reply_text(f"Пользователь @{target} не имеет задач.")
        return
    period = "за всё время" if not days else f"за последние {days} дней"
    response = f"📊 Статистика @{target} {period}:\n"
    response += f"🐛 Баги: {stats['bug']['total']} создано, {stats['bug']['closed']} закрыто\n"
    response += f"💡 Предложения: {stats['suggestion']['total']} создано, {stats['suggestion']['closed']} закрыто"
    await update.message.reply_text(response)

async def tags_list_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tags = get_all_tags()
    if not tags:
        await update.message.reply_text("📭 Нет тегов.")
        return
    await update.message.reply_text("🏷️ Теги:\n" + "\n".join(tags))

async def tag_stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Укажите тег: /tagstats #модуль")
        return
    tag = context.args[0].strip()
    if not tag.startswith('#'):
        tag = '#' + tag
    days = None
    if len(context.args) > 1 and context.args[1].lower() in ["week", "7"]:
        days = 7
    elif len(context.args) > 1 and context.args[1].lower() in ["month", "30"]:
        days = 30
    stats = get_stats_by_tag(tag, days)
    if stats["bug"]["total"] == 0 and stats["suggestion"]["total"] == 0:
        await update.message.reply_text(f"По тегу {tag} задач нет.")
        return
    period = "за всё время" if not days else f"за последние {days} дней"
    response = f"📊 Статистика по тегу {tag} {period}:\n"
    response += f"🐛 Баги: {stats['bug']['total']} создано, {stats['bug']['closed']} закрыто\n"
    response += f"💡 Предложения: {stats['suggestion']['total']} создано, {stats['suggestion']['closed']} закрыто"
    await update.message.reply_text(response)

async def open_tasks_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    priority = None
    tag = None
    days = None
    for arg in args:
        if arg.startswith('#'):
            tag = arg
        elif arg.lower() in ['high', 'medium', 'low']:
            priority = arg.lower()
        elif arg.lower() in ['today', 'day']:
            days = 1
        elif arg.lower() in ['week', '7']:
            days = 7
        elif arg.lower() in ['month', '30']:
            days = 30
    rows = get_open_tasks_text(limit=10, priority=priority, tag=tag, days=days)
    if not rows:
        await update.message.reply_text("📭 Нет открытых задач по заданным фильтрам.")
        return
    response = "📋 Открытые задачи:\n"
    for row in rows:
        issue_id, author, text, issue_type, priority, tags, created = row
        emoji = {"high": "🔴", "medium": "🟡", "low": "🟢"}.get(priority, "")
        created_str = created[:16] if isinstance(created, str) else created.strftime("%Y-%m-%d %H:%M")
        response += f"#{issue_id} {issue_type} {emoji} от {author} ({created_str}): {text[:50]}...\n"
    await update.message.reply_text(response)

async def joke_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"😂 {get_random_joke()}")

async def close_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Укажите ID: /close 3")
        return
    try:
        issue_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Некорректный ID.")
        return
    issue = get_issue_by_id(issue_id)
    if not issue:
        await update.message.reply_text(f"Задача #{issue_id} не найдена.")
        return
    if issue[1] == 'closed':
        await update.message.reply_text(f"Задача #{issue_id} уже закрыта.")
        return
    close_issue(issue_id, update.effective_user.id)
    await update.message.reply_text(f"✅ Задача #{issue_id} закрыта.")

async def reopen_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Укажите ID: /reopen 3")
        return
    try:
        issue_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Некорректный ID.")
        return
    issue = get_issue_by_id(issue_id)
    if not issue:
        await update.message.reply_text(f"Задача #{issue_id} не найдена.")
        return
    if issue[1] == 'open':
        await update.message.reply_text(f"Задача #{issue_id} уже открыта.")
        return
    reopen_issue(issue_id, update.effective_user.id)
    await update.message.reply_text(f"🔄 Задача #{issue_id} открыта заново.")

async def comment_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text("Укажите ID и текст: /comment 3 текст")
        return
    try:
        issue_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("Некорректный ID.")
        return
    text = " ".join(context.args[1:])
    issue = get_issue_by_id(issue_id)
    if not issue:
        await update.message.reply_text(f"Задача #{issue_id} не найдена.")
        return
    user = update.effective_user
    add_comment(issue_id, user.id, user.full_name or user.username, text)
    await update.message.reply_text(f"💬 Комментарий к задаче #{issue_id} добавлен.")

async def find_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Укажите текст: /find текст")
        return
    query = " ".join(context.args)
    rows = search_issues(query)
    if not rows:
        await update.message.reply_text("Ничего не найдено.")
        return
    response = "🔍 Результаты:\n"
    for row in rows:
        issue_id, author, text, issue_type, status, tags, created = row
        emoji = "✅" if status == "closed" else "❌"
        created_str = created[:16] if isinstance(created, str) else created.strftime("%Y-%m-%d %H:%M")
        tag_display = f" {tags}" if tags else ""
        response += f"#{issue_id} {issue_type} {emoji} от {author} ({created_str}): {text[:50]}...{tag_display}\n"
    await update.message.reply_text(response)

async def stats_responsible_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Укажите юзернейм: /stats_responsible @username")
        return
    target = context.args[0].lstrip('@')
    days = None
    if len(context.args) > 1 and context.args[1].lower() in ["week", "7"]:
        days = 7
    elif len(context.args) > 1 and context.args[1].lower() in ["month", "30"]:
        days = 30
    total, closed, bugs, suggestions = get_stats_responsible(target, days)
    period = "за всё время" if not days else f"за последние {days} дней"
    response = f"📊 Статистика по ответственному @{target} {period}:\n"
    response += f"📌 Всего: {total}\n✅ Закрыто: {closed}\n🐛 Багов: {bugs}\n💡 Предложений: {suggestions}"
    await update.message.reply_text(response)

async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    days = None
    if context.args and context.args[0].isdigit():
        days = int(context.args[0])
    rows = generate_export(days)
    if not rows:
        await update.message.reply_text("Нет данных.")
        return
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Автор", "Юзернейм", "Текст", "Тип", "Статус", "Приоритет", "Теги", "Ответственные", "Создано"])
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    await update.message.reply_document(
        document=output.getvalue().encode('utf-8'),
        filename=f"tasks_export_{datetime.now().strftime('%Y%m%d')}.csv",
        caption="📊 Выгрузка"
    )

async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_created, total_closed, type_counts, priority_counts = generate_weekly_report()
    response = f"📈 Отчёт за 7 дней:\n📌 Создано: {total_created}\n✅ Закрыто: {total_closed}\n\nПо типам:\n"
    for t, cnt in type_counts:
        response += f"  {t}: {cnt}\n"
    response += "\nПо приоритетам:\n"
    priority_names = {'high': '🔴 Критичный', 'medium': '🟡 Важный', 'low': '🟢 Обычный'}
    for p, cnt in priority_counts:
        response += f"  {priority_names.get(p, p)}: {cnt}\n"
    await update.message.reply_text(response)

async def report_pdf_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    total_created, total_closed, type_counts, priority_counts = generate_weekly_report()
    buffer = io_bytes.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.drawString(100, height-50, "Еженедельный отчёт по задачам")
    c.drawString(100, height-80, f"Создано за неделю: {total_created}")
    c.drawString(100, height-100, f"Закрыто за неделю: {total_closed}")
    y = height - 130
    c.drawString(100, y, "По типам:")
    y -= 20
    for t, cnt in type_counts:
        c.drawString(120, y, f"{t}: {cnt}")
        y -= 20
    y -= 10
    c.drawString(100, y, "По приоритетам:")
    y -= 20
    for p, cnt in priority_counts:
        c.drawString(120, y, f"{p}: {cnt}")
        y -= 20
    c.save()
    buffer.seek(0)
    await update.message.reply_document(
        document=InputFile(buffer, filename="weekly_report.pdf"),
        caption="📄 Еженедельный отчёт в PDF"
    )

async def report_excel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    from_date = None
    to_date = None
    for i, arg in enumerate(args):
        if arg.lower() == 'from' and i+1 < len(args):
            try:
                from_date = datetime.strptime(args[i+1], "%Y-%m-%d").strftime("%Y-%m-%d %H:%M:%S")
            except:
                pass
        elif arg.lower() == 'to' and i+1 < len(args):
            try:
                to_date = datetime.strptime(args[i+1], "%Y-%m-%d").strftime("%Y-%m-%d %H:%M:%S")
            except:
                pass
    if not from_date and not to_date:
        from_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")
        to_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    await update.message.reply_text("📊 Генерирую Excel-отчёт...")
    try:
        excel_file = generate_excel_report(from_date, to_date)
        filename = f"report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        await update.message.reply_document(
            document=excel_file,
            filename=filename,
            caption=f"📋 Отчёт по закрытым задачам с {from_date[:10] if from_date else 'начала'} по {to_date[:10] if to_date else 'конец'}"
        )
    except Exception as e:
        logger.error(f"Ошибка генерации Excel-отчёта: {e}")
        await update.message.reply_text("❌ Не удалось сгенерировать отчёт.")

async def rating_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    username = update.effective_user.username or "без юзернейма"
    points = get_points(user_id)
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM rating WHERE points > ?", (points,))
    rank = c.fetchone()[0] + 1
    conn.close()
    await update.message.reply_text(
        f"📊 Ваш рейтинг:\n"
        f"Пользователь: @{username}\n"
        f"Очки: {points}\n"
        f"Позиция в топе: {rank}"
    )

async def top_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    top_users = get_top_users(10)
    if not top_users:
        await update.message.reply_text("Пока нет рейтинга.")
        return
    response = "🏆 Топ-10 пользователей:\n"
    for i, (uid, pts) in enumerate(top_users, 1):
        conn = sqlite3.connect("issues.db")
        c = conn.cursor()
        c.execute("SELECT author_name FROM issues WHERE author_id=? LIMIT 1", (uid,))
        row = c.fetchone()
        conn.close()
        name = row[0] if row else f"user_{uid}"
        response += f"{i}. {name} – {pts} очков\n"
    await update.message.reply_text(response)

async def dashboard_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    conn = sqlite3.connect("issues.db")
    c = conn.cursor()
    week_ago = datetime.now() - timedelta(days=7)
    c.execute("SELECT COUNT(*) FROM issues WHERE created_at >= ?", (week_ago,))
    total_created = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM issues WHERE status='closed' AND created_at >= ?", (week_ago,))
    total_closed = c.fetchone()[0]
    c.execute("SELECT priority, COUNT(*) FROM issues WHERE created_at >= ? GROUP BY priority", (week_ago,))
    priority_counts = c.fetchall()
    c.execute("SELECT tags FROM issues WHERE tags IS NOT NULL AND tags != '' ORDER BY created_at DESC LIMIT 50")
    tags_rows = c.fetchall()
    tag_freq = {}
    for row in tags_rows:
        if row[0]:
            for tag in row[0].split(','):
                tag = tag.strip()
                tag_freq[tag] = tag_freq.get(tag, 0) + 1
    top_tags = sorted(tag_freq.items(), key=lambda x: x[1], reverse=True)[:5]
    top_users = get_top_users(5)
    conn.close()

    html = f"""
    <html>
    <head><meta charset="utf-8"><title>Дашборд</title></head>
    <body>
    <h1>📊 Дашборд задач</h1>
    <p><b>Создано за неделю:</b> {total_created}</p>
    <p><b>Закрыто за неделю:</b> {total_closed}</p>
    <h2>Приоритеты</h2>
    <ul>
    """
    for p, cnt in priority_counts:
        html += f"<li>{p}: {cnt}</li>"
    html += "</ul><h2>Топ-5 тегов</h2><ul>"
    for tag, cnt in top_tags:
        html += f"<li>{tag}: {cnt}</li>"
    html += "</ul><h2>Топ-5 пользователей</h2><ul>"
    for uid, pts in top_users:
        html += f"<li>ID {uid}: {pts} очков</li>"
    html += "</ul></body></html>"

    html_bytes = html.encode('utf-8')
    await update.message.reply_document(
        document=io.BytesIO(html_bytes),
        filename="dashboard.html",
        caption="📈 Дашборд (откройте в браузере)"
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = """
🤖 Доступные команды:

/bug <описание> – создать баг
/suggest <описание> – создать предложение
/ask <вопрос> – задать вопрос ИИ
/mystats – ваша статистика
/stats @user – статистика пользователя (админ)
/tags – все теги
/tagstats #тег – статистика по тегу
/open [high|medium|low] [#тег] [today|week|month] – открытые задачи с фильтрами
/close <id> – закрыть задачу
/reopen <id> – открыть заново
/comment <id> текст – комментарий
/find текст – поиск
/priority <id> <high|medium|low> – изменить приоритет
/stats_responsible @user – статистика по ответственному
/export – выгрузить CSV
/report – отчёт за неделю
/report_pdf – отчёт в PDF
/report_excel [from YYYY-MM-DD] [to YYYY-MM-DD] – отчёт в Excel по закрытым задачам
/joke – анекдот
/rating – ваш рейтинг
/top – топ-10 пользователей
/dashboard – дашборд (HTML)

Админские команды:
/add_responsible @user – добавить ответственного
/remove_responsible @user – удалить ответственного
/list_responsible – список ответственных
/set_auto_close <дни> – автозакрытие через N дней
/ban_user <id> – забанить пользователя
/unban_user <id> – разбанить
/help – это сообщение
"""
    await update.message.reply_text(help_text)

# ---------- КОМАНДЫ АДМИНА ----------
async def add_responsible_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Укажите юзернейм: /add_responsible @username")
        return
    username = context.args[0].lstrip('@')
    add_responsible(username)
    await update.message.reply_text(f"✅ @{username} добавлен.")

async def remove_responsible_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Укажите юзернейм: /remove_responsible @username")
        return
    username = context.args[0].lstrip('@')
    remove_responsible(username)
    await update.message.reply_text(f"✅ @{username} удалён.")

async def list_responsible_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    list_resp = get_responsible_list()
    if not list_resp:
        await update.message.reply_text("Список ответственных пуст.")
        return
    response = "📋 Список ответственных:\n" + "\n".join([f"@{u}" for u in list_resp])
    await update.message.reply_text(response)

async def set_auto_close_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Укажите количество дней: /set_auto_close 14")
        return
    try:
        days = int(context.args[0])
        set_setting("auto_close_days", str(days))
        await update.message.reply_text(f"✅ Автозакрытие установлено на {days} дней.")
    except ValueError:
        await update.message.reply_text("Некорректное число.")

async def ban_user_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Укажите ID пользователя: /ban_user 123456789")
        return
    try:
        user_id = int(context.args[0])
        ban_user(user_id)
        await update.message.reply_text(f"✅ Пользователь {user_id} забанен.")
    except ValueError:
        await update.message.reply_text("Некорректный ID.")

async def unban_user_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет прав.")
        return
    if not context.args:
        await update.message.reply_text("Укажите ID пользователя: /unban_user 123456789")
        return
    try:
        user_id = int(context.args[0])
        unban_user(user_id)
        await update.message.reply_text(f"✅ Пользователь {user_id} разбанен.")
    except ValueError:
        await update.message.reply_text("Некорректный ID.")

# ---------- ЗАПУСК ----------
def main():
    init_db()
    app = Application.builder().token(TOKEN).build()
    if app.job_queue is None:
        logger.warning("JobQueue не инициализирован")
    if app.bot_data.get('reminder_jobs') is None:
        app.bot_data['reminder_jobs'] = {}

    app.add_handler(MessageHandler(filters.StatusUpdate.NEW_CHAT_MEMBERS, greet_new_member))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CallbackQueryHandler(priority_callback, pattern=r"^(set_priority_\d+_(high|medium|low)|skip_priority_\d+)$"))
    app.add_handler(CallbackQueryHandler(confirm_callback, pattern="^(confirm_yes|confirm_no)$"))
    app.add_handler(CallbackQueryHandler(responsible_callback, pattern=r"^(resp_.+|resp_skip|resp_other)$"))
    app.add_handler(CallbackQueryHandler(vote_callback, pattern=r"^vote_\d+_[-\d]+$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_manual_responsible))

    app.add_handler(CommandHandler("bug", bug_command))
    app.add_handler(CommandHandler("suggest", suggest_command))
    app.add_handler(CommandHandler("ask", ask_command))
    app.add_handler(CommandHandler("priority", priority_command))
    app.add_handler(CommandHandler("mystats", my_stats_command))
    app.add_handler(CommandHandler("stats", stats_username_command))
    app.add_handler(CommandHandler("tags", tags_list_command))
    app.add_handler(CommandHandler("tagstats", tag_stats_command))
    app.add_handler(CommandHandler("open", open_tasks_command))
    app.add_handler(CommandHandler("joke", joke_command))
    app.add_handler(CommandHandler("close", close_command))
    app.add_handler(CommandHandler("reopen", reopen_command))
    app.add_handler(CommandHandler("comment", comment_command))
    app.add_handler(CommandHandler("find", find_command))
    app.add_handler(CommandHandler("stats_responsible", stats_responsible_command))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("report", report_command))
    app.add_handler(CommandHandler("report_pdf", report_pdf_command))
    app.add_handler(CommandHandler("report_excel", report_excel_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("rating", rating_command))
    app.add_handler(CommandHandler("top", top_command))
    app.add_handler(CommandHandler("dashboard", dashboard_command))

    app.add_handler(CommandHandler("add_responsible", add_responsible_command))
    app.add_handler(CommandHandler("remove_responsible", remove_responsible_command))
    app.add_handler(CommandHandler("list_responsible", list_responsible_command))
    app.add_handler(CommandHandler("set_auto_close", set_auto_close_command))
    app.add_handler(CommandHandler("ban_user", ban_user_command))
    app.add_handler(CommandHandler("unban_user", unban_user_command))

    async def delete_webhook():
        await app.bot.delete_webhook()
        logger.info("Webhook удалён")

    loop = asyncio.get_event_loop()
    loop.run_until_complete(delete_webhook())

    if app.job_queue:
        try:
            morning_time = datetime.strptime(MORNING_TIME_UTC, "%H:%M").time()
            app.job_queue.run_daily(morning_greeting, time=morning_time, days=tuple(range(7)))
        except Exception as e:
            logger.error(f"Ошибка планирования: {e}")

    logger.info("Бот с расширенными функциями запущен!")
    app.run_polling()

if __name__ == "__main__":
    main()
